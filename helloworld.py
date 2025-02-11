from PIL import Image
import openpyxl
from openpyxl.styles import PatternFill

def rgb_to_hex(rgb):
    return '{:02x}{:02x}{:02x}'.format(rgb[0], rgb[1], rgb[2])

def image_to_excel(image_path, output_excel):
    # Load the image
    img = Image.open(image_path)
    img = img.convert("RGB")

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    max_size = 180
    img.thumbnail((max_size, max_size))

    width, height = img.size

    for y in range(height):
        for x in range(width):
            r, g, b = img.getpixel((x, y))
            hex_color = rgb_to_hex((r, g, b))

            cell = sheet.cell(row=y + 1, column=x + 1)

            fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
            cell.fill = fill

            sheet.column_dimensions[openpyxl.utils.get_column_letter(x + 1)].width = 3.0
        sheet.row_dimensions[y + 1].height = 25

    workbook.save(output_excel)


image_path = r''  # Path to your image inside ''
output_excel = ''  # Path to save the Excel file inside ''
image_to_excel(image_path, output_excel)
